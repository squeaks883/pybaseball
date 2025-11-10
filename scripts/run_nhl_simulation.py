"""Utility for running NHL Monte Carlo simulations against Elo ratings.

This module replicates the workflow described in the Hybrid Ice v2.3
specification.  It loads several Excel / CSV inputs, runs Monte Carlo
simulations for user supplied matchups, prints the resulting win
probabilities to the console, and stores the output back into the master
workbook.

The script intentionally keeps the data wrangling light weight so that it
can function even when a subset of the supporting workbooks is present.
"""
from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook


@dataclass(frozen=True)
class Matchup:
    """Simple representation of an NHL game to simulate."""

    away: str
    home: str

    @classmethod
    def parse(cls, token: str) -> "Matchup":
        try:
            away, home = token.split("@")
        except ValueError as exc:  # pragma: no cover - defensive guard
            raise argparse.ArgumentTypeError(
                "Matchups must use the 'AWAY@HOME' format"
            ) from exc
        return cls(away.strip().upper(), home.strip().upper())

    def label(self) -> str:
        return f"{self.away} @ {self.home}"


def load_elo_table(path: Path) -> pd.DataFrame:
    """Load the Elo table expected by the simulation engine."""

    df = pd.read_csv(path)
    if {"Team", "Elo", "HFA"}.difference(df.columns):  # pragma: no cover - IO guard
        missing = ", ".join(sorted({"Team", "Elo", "HFA"}.difference(df.columns)))
        raise ValueError(
            f"Elo file '{path}' is missing required columns: {missing}"
        )
    return df


def simulate_matchup(matchup: Matchup, elo_table: pd.DataFrame, sims: int, rng: np.random.Generator) -> dict[str, float]:
    """Run a Monte Carlo simulation for a single matchup."""

    home_row = elo_table.loc[elo_table["Team"] == matchup.home]
    away_row = elo_table.loc[elo_table["Team"] == matchup.away]
    if home_row.empty or away_row.empty:
        missing = matchup.home if home_row.empty else matchup.away
        raise ValueError(f"Missing Elo information for team '{missing}'")

    home_elo = home_row["Elo"].iloc[0]
    home_adv = home_row["HFA"].iloc[0]
    away_elo = away_row["Elo"].iloc[0]

    expected_home = 1.0 / (1.0 + 10 ** ((away_elo - (home_elo + home_adv)) / 400.0))
    sims_outcome = rng.binomial(1, expected_home, sims)
    home_win = float(np.mean(sims_outcome))
    away_win = 1.0 - home_win

    return {
        "Matchup": matchup.label(),
        "Home_Win%": round(home_win * 100, 2),
        "Away_Win%": round(away_win * 100, 2),
        "Edge": round((home_win - 0.5) * 100, 2),
    }


def save_results(
    workbook_path: Path, sheet_name: str, results: Sequence[dict[str, float]], timestamp: dt.datetime
) -> None:
    """Append results to the workbook, creating the sheet when necessary."""

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    if ws.max_row == 1 and all(cell.value is None for cell in ws[1]):  # pragma: no cover - IO guard
        ws.append(["Matchup", "Home_Win%", "Away_Win%", "Edge", "Timestamp"])

    for result in results:
        ws.append([
            result["Matchup"],
            result["Home_Win%"],
            result["Away_Win%"],
            result["Edge"],
            timestamp.isoformat(timespec="seconds"),
        ])
    wb.save(workbook_path)


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run Hybrid Ice NHL simulations")
    parser.add_argument("--date", type=dt.date.fromisoformat, required=True, help="Simulation date (YYYY-MM-DD)")
    parser.add_argument("--matchup", type=Matchup.parse, action="append", dest="matchups", required=True)
    parser.add_argument("--sims", type=int, default=10_000, help="Number of Monte Carlo runs (default: 10000)")
    parser.add_argument("--elo", type=Path, default=Path("NHL_Team_Home_Ice_Elo.csv"))
    parser.add_argument("--master", type=Path, default=Path("master - nhl.xlsx"))
    parser.add_argument("--seed", type=int, default=None, help="Optional random seed for reproducibility")
    return parser.parse_args(list(argv) if argv is not None else None)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)
    elo_table = load_elo_table(args.elo)

    rng = np.random.default_rng(args.seed)
    results = [simulate_matchup(matchup, elo_table, args.sims, rng) for matchup in args.matchups]

    output_df = pd.DataFrame(results)
    print("\n🏒 Simulation Results:")
    print(output_df.to_string(index=False))

    sheet_name = f"Simulation_{args.date.isoformat()}"
    save_results(args.master, sheet_name, results, dt.datetime.now())

    print(f"\n💾 Saved simulation results → {args.master}")
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())
