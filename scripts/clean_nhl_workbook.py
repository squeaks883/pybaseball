"""Utility to remove `meta_` prefixed worksheets from an Excel workbook.

The cleaner is intentionally conservative and will refuse to run when the
workbook cannot be located.  When a workbook is found, the script removes any
worksheet whose name begins with ``meta_`` (case insensitive) and saves the
workbook under the original file name.  The tool prints a concise summary of
its actions so it can be used as part of larger automation pipelines.
"""
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, List

from openpyxl import load_workbook


def _collect_meta_sheets(sheetnames: Iterable[str]) -> List[str]:
    """Return a list of worksheets whose name starts with ``meta_``.

    The comparison is performed in a case-insensitive manner so that sheets such
    as ``META_data`` or ``Meta_summary`` are detected.
    """

    return [name for name in sheetnames if name.lower().startswith("meta_")]


def clean_workbook(path: Path, dry_run: bool = False) -> List[str]:
    """Remove ``meta_`` sheets from ``path`` and return the removed sheet names.

    Parameters
    ----------
    path:
        Path to the workbook to clean.
    dry_run:
        When ``True`` the workbook is inspected and the list of matching sheets
        is returned without mutating the file.
    """

    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    workbook = load_workbook(path)
    sheets_to_remove = _collect_meta_sheets(workbook.sheetnames)

    if not dry_run:
        for sheet in sheets_to_remove:
            del workbook[sheet]
        workbook.save(path)

    return sheets_to_remove


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Clean meta_ sheets from a workbook")
    parser.add_argument(
        "path",
        type=Path,
        help="Path to the workbook (e.g. 'master - nhl 2.xlsx')",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Report sheets that would be removed without modifying the workbook.",
    )
    return parser


def main() -> int:
    parser = _build_parser()
    args = parser.parse_args()

    try:
        removed_sheets = clean_workbook(args.path, dry_run=args.dry_run)
    except FileNotFoundError as exc:
        parser.error(str(exc))
        return 2

    if not removed_sheets:
        print("No meta_ sheets found. Workbook left unchanged.")
    else:
        if args.dry_run:
            print("Sheets that would be removed:")
        else:
            print("Removed sheets:")
        for sheet in removed_sheets:
            print(f"  - {sheet}")
        if not args.dry_run:
            print("Workbook saved with meta_ sheets removed.")

    return 0


if __name__ == "__main__":  # pragma: no cover - manual invocation entry-point
    raise SystemExit(main())
