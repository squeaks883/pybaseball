"""Hybrid Ice v1.2 Simulation for Chicago Blackhawks @ Vancouver Canucks.

This script replicates the provided Hybrid Ice v1.2 workflow to:
    * ensure the team worksheets exist in the master workbook,
    * populate them with the supplied lineups and real ratings,
    * calculate composite team ratings from the top skaters,
    * update Home Ice Elo values and their bulk log CSV,
    * run a Poisson-based Monte Carlo simulation,
    * print a summary of the results, and
    * export the main outputs to a CSV file.

Paths are relative to the working directory. Running the script will mutate the
Excel workbook and the Elo CSV/log files, and write a summary CSV of the latest
simulation run.
"""

from __future__ import annotations

import datetime as dt
import json
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


MASTER_PATH = Path("master - nhl.xlsx")
ELO_CSV_PATH = Path("NHL_Team_Home_Ice_Elo.csv")
ELO_LOG_CSV_PATH = Path("NHL_Bulk_Elo_Log_teamHFA.csv")
OUTPUT_CSV_PATH = Path("Hybrid_Ice_Results_Chicago_Vancouver.csv")


def timestamp() -> str:
    """Return a formatted timestamp for logging."""

    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def find_col(headers: list[str], patterns: list[str]) -> int | None:
    """Locate the index of the first header that matches any regex pattern."""

    for idx, value in enumerate(headers):
        if value and any(re.search(pattern, str(value).lower()) for pattern in patterns):
            return idx
    return None


def create_team_sheet(wb, team: str, players: list[tuple[str, str, int, float]], stamp: str) -> None:
    """Ensure the team sheet exists and is populated with player data."""

    if team in wb.sheetnames:
        return

    ws = wb.create_sheet(team)
    headers = [
        "Player",
        "Position",
        "Line",
        "Real_Rating",
        "G",
        "A",
        "PTS",
        "TOI",
        "Last_Updated",
    ]
    ws.append(headers)
    for player in players:
        ws.append(list(player) + [None, None, None, None, stamp])


def load_player_ratings(wb, teams: list[str]) -> pd.DataFrame:
    """Extract player real ratings from the provided worksheets."""

    player_ratings: list[dict[str, object]] = []

    for team in teams:
        if team not in wb.sheetnames:
            continue

        ws = wb[team]
        header_row = [cell.value for cell in ws[1]]
        name_col = find_col(header_row, [r"name", r"player"])
        rr_col = find_col(header_row, [r"real", r"rr", r"rating"])
        if name_col is None or rr_col is None:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_col]
            real_rating = row[rr_col]
            if name and isinstance(real_rating, (int, float)):
                player_ratings.append({"team": team, "player": name, "real_rating": real_rating})

    return pd.DataFrame(player_ratings)


def compute_composite_ratings(df: pd.DataFrame) -> dict[str, float]:
    """Calculate the composite team rating (average of the top nine ratings)."""

    return (
        df.groupby("team")["real_rating"].apply(lambda series: np.mean(sorted(series, reverse=True)[:9])).to_dict()
    )


def update_elo(composite: dict[str, float], stamp: str) -> None:
    """Update the Home Ice Elo CSV and append to the log CSV."""

    elo = pd.read_csv(ELO_CSV_PATH)
    log = pd.read_csv(ELO_LOG_CSV_PATH)

    if not composite:
        return

    average_rating = np.mean(list(composite.values()))

    for team, value in composite.items():
        team_key = team.split()[0]
        mask = elo["Home"].str.contains(team_key, case=False, na=False)
        if mask.any():
            new_elo = 1500 + 25 * (value - average_rating)
            elo.loc[mask, "Home_Ice_Elo"] = new_elo
            log.loc[len(log)] = {"timestamp": stamp, "team": team, "elo": new_elo}

    elo.to_csv(ELO_CSV_PATH, index=False)
    log.to_csv(ELO_LOG_CSV_PATH, index=False)


def run_simulation(composite: dict[str, float]) -> dict[str, float | dict[str, float]]:
    """Execute the Poisson simulation and return the summary statistics."""

    np.random.seed(42)

    chicago_rating = round(composite.get("Chicago Blackhawks", 3.0), 3)
    vancouver_rating = round(composite.get("Vancouver Canucks", 3.0), 3)

    chicago_expectation = 2.9 + (chicago_rating - vancouver_rating) * 0.15
    vancouver_expectation = 2.9 + (vancouver_rating - chicago_rating) * 0.15

    chicago_goals = np.random.poisson(max(0.5, chicago_expectation), 10_000)
    vancouver_goals = np.random.poisson(max(0.5, vancouver_expectation), 10_000)

    chicago_win = np.mean(chicago_goals > vancouver_goals)
    vancouver_win = np.mean(vancouver_goals > chicago_goals)
    tie_probability = np.mean(chicago_goals == vancouver_goals)
    mean_total = (chicago_goals + vancouver_goals).mean()

    return {
        "chicago_composite": chicago_rating,
        "vancouver_composite": vancouver_rating,
        "chicago_win_pct": round(chicago_win, 3),
        "vancouver_win_pct": round(vancouver_win, 3),
        "tie_pct": round(tie_probability, 3),
        "mean_total": round(mean_total, 2),
    }


def enrich_player_outputs(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, float]]:
    """Calculate player-level probabilities and team real rating totals."""

    df = df.copy()
    df["prob_point"] = (0.12 * df["real_rating"]).clip(0, 0.65)
    df["prob_2plus"] = (0.03 * df["real_rating"]).clip(0, 0.35)

    top_five = df.sort_values("real_rating", ascending=False).head(5)[["team", "player", "real_rating"]]
    three_leg = df.sort_values("prob_point", ascending=False).head(3)[["team", "player", "prob_point"]]
    two_plus = df.sort_values("prob_2plus", ascending=False).head(2)[["team", "player", "prob_2plus"]]
    assist_parlay = df.sort_values("real_rating", ascending=False).head(3)[["team", "player", "real_rating"]]

    team_totals = df.groupby("team")["real_rating"].apply(lambda series: series.nlargest(6).sum()).to_dict()

    return top_five, three_leg, two_plus, assist_parlay, team_totals


def export_summary(summary: dict[str, float], mean_total: float) -> None:
    """Write the summary CSV with composite ratings and win probabilities."""

    summary_df = pd.DataFrame(
        {
            "Team": ["Chicago Blackhawks", "Vancouver Canucks"],
            "Composite_Rating": [summary["chicago_composite"], summary["vancouver_composite"]],
            "Win_Pct": [summary["chicago_win_pct"], summary["vancouver_win_pct"]],
            "Tie_Pct": [summary["tie_pct"], summary["tie_pct"]],
            "Mean_Total_Goals": [mean_total, mean_total],
        }
    )

    summary_df.to_csv(OUTPUT_CSV_PATH, index=False)


def print_summary(stamp: str, summary: dict[str, float], top_five, three_leg, two_plus, assist_parlay, team_totals) -> None:
    """Print the Hybrid Ice summary and parlays to stdout."""

    print("\n===== HYBRID ICE v1.2 SIMULATION SUMMARY =====\n")
    print(f"Date: {stamp}")
    print(f"Chicago Composite Rating: {summary['chicago_composite']}")
    print(f"Vancouver Composite Rating: {summary['vancouver_composite']}")
    print(f"Chicago Win %: {summary['chicago_win_pct'] * 100:.1f}%")
    print(f"Vancouver Win %: {summary['vancouver_win_pct'] * 100:.1f}%")
    print(f"Tie Probability: {summary['tie_pct'] * 100:.1f}%")
    print(f"Model Mean Total Goals: {summary['mean_total']}\n")
    print("Top 5 Real Ratings:")
    print(top_five.to_string(index=False))
    print("\n3-Leg Point Parlay:")
    print(three_leg.to_string(index=False))
    print("\n2 Players 2+ Points:")
    print(two_plus.to_string(index=False))
    print("\n3-Leg Assists Parlay:")
    print(assist_parlay.to_string(index=False))
    print("\nTeam Real Rating Totals:")
    print(json.dumps(team_totals, indent=2))
    print(f"\nResults exported to {OUTPUT_CSV_PATH}")


def main() -> None:
    stamp = timestamp()

    wb = load_workbook(MASTER_PATH)

    teams = ["Chicago Blackhawks", "Vancouver Canucks"]

    chicago_players = [
        ("Ryan Greene", "LW", 1, 5.0),
        ("Connor Bedard", "C", 1, 5.0),
        ("Andre Burakovsky", "RW", 1, 5.0),
        ("Teuvo Teravainen", "LW", 2, 4.2),
        ("Frank Nazar", "C", 2, 4.2),
        ("Tyler Bertuzzi", "RW", 2, 4.2),
        ("Ryan Donato", "LW", 3, 3.4),
        ("Colton Dach", "C", 3, 3.4),
        ("Ilya Mikheyev", "RW", 3, 3.4),
        ("Oliver Moore", "LW", 4, 2.2),
        ("Nick Foligno", "C", 4, 2.2),
        ("Sam Lafferty", "RW", 4, 2.0),
        ("Alex Vlasic", "D", 1, 4.4),
        ("Sam Rinzel", "D", 1, 4.4),
        ("Matt Grzelcyk", "D", 2, 3.8),
        ("Connor Murphy", "D", 2, 3.8),
        ("Wyatt Kaiser", "D", 3, 3.2),
        ("Artyom Levshunov", "D", 3, 3.2),
        ("Spencer Knight", "G", 0, 4.0),
    ]

    vancouver_players = [
        ("Evander Kane", "LW", 1, 5.0),
        ("Elias Pettersson", "C", 1, 5.0),
        ("Conor Garland", "RW", 1, 5.0),
        ("Jake DeBrusk", "LW", 2, 4.2),
        ("Lukas Reichel", "C", 2, 4.2),
        ("Brock Boeser", "RW", 2, 4.2),
        ("Drew O'Connor", "LW", 3, 3.4),
        ("Aatu Raty", "C", 3, 3.4),
        ("Kiefer Sherwood", "RW", 3, 3.4),
        ("Arshdeep Bains", "LW", 4, 2.2),
        ("Max Sasson", "C", 4, 2.2),
        ("Linus Karlsson", "RW", 4, 2.2),
        ("Quinn Hughes", "D", 1, 4.4),
        ("Tyler Myers", "D", 1, 4.4),
        ("Marcus Pettersson", "D", 2, 3.8),
        ("Filip Hronek", "D", 2, 3.8),
        ("Elias Nils Pettersson", "D", 3, 3.2),
        ("Tom Willander", "D", 3, 3.2),
        ("Kevin Lankinen", "G", 0, 4.0),
    ]

    create_team_sheet(wb, "Chicago Blackhawks", chicago_players, stamp)
    create_team_sheet(wb, "Vancouver Canucks", vancouver_players, stamp)
    wb.save(MASTER_PATH)

    player_df = load_player_ratings(wb, teams)

    composite_ratings = compute_composite_ratings(player_df)
    update_elo(composite_ratings, stamp)

    summary = run_simulation(composite_ratings)
    top_five, three_leg, two_plus, assist_parlay, team_totals = enrich_player_outputs(player_df)

    results = {
        "timestamp": stamp,
        **summary,
        "team_real_rating_totals": team_totals,
    }

    export_summary(summary, results["mean_total"])
    print_summary(stamp, results, top_five, three_leg, two_plus, assist_parlay, team_totals)


if __name__ == "__main__":
    main()

